import csv
import os
import json
import tempfile
from datetime import datetime, date
from flask import Flask, render_template, jsonify, request, send_file
from openpyxl import load_workbook
import xlrd

app = Flask(__name__)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'data')
SNAPSHOT_DIR = os.path.join(DATA_DIR, 'snapshots')
os.makedirs(SNAPSHOT_DIR, exist_ok=True)

PRODUCTS_CSV = os.path.join(DATA_DIR, 'products.csv')
INVENTORY_CSV = os.path.join(DATA_DIR, 'inventory.csv')
CONTAINERS_CSV = os.path.join(DATA_DIR, 'containers.csv')
LAST_IMPORT_JSON = os.path.join(DATA_DIR, 'last_import.json')

WEEKS_IN_YEAR = 52
WARN_WEEKS = 4


def get_weeks_elapsed():
    last = get_last_import()
    if last.get('date') and last['date'] != 'N/A':
        try:
            d = datetime.strptime(last['date'], '%Y-%m-%d').date()
            sales_year = 2025
            if d.year == sales_year:
                return max(1, int((d - date(sales_year, 1, 1)).days / 7))
            return 52
        except (ValueError, TypeError):
            pass
    return 33


def read_csv(path):
    if not os.path.exists(path):
        return []
    with open(path, 'r', encoding='utf-8-sig') as f:
        return list(csv.DictReader(f))


def write_csv(path, rows, fieldnames):
    with open(path, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(rows)


def safe_float(val, default=0):
    if val is None or val == '':
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def get_last_import():
    if os.path.exists(LAST_IMPORT_JSON):
        with open(LAST_IMPORT_JSON, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {'date': 'N/A', 'filename': 'N/A', 'item_count': 0}


def calculate_status(j, sales_2025, p_diff, i_intransit, weeks_elapsed):
    if j <= 0:
        return 'red'
    weekly_rate = sales_2025 / weeks_elapsed if weeks_elapsed > 0 and sales_2025 > 0 else 0
    if weekly_rate > 0:
        weeks_left = j / weekly_rate
        if weeks_left < WARN_WEEKS:
            return 'yellow'
    if p_diff is not None and p_diff < 0 and weekly_rate == 0:
        return 'yellow'
    return 'green'


def get_merged_data():
    products = read_csv(PRODUCTS_CSV)
    inventory = {r['item_code']: r for r in read_csv(INVENTORY_CSV)}
    weeks_elapsed = get_weeks_elapsed()

    merged = []
    for p in products:
        code = p['item_code']
        inv = inventory.get(code, {})
        g = safe_float(inv.get('g_inventory'))
        h = safe_float(inv.get('h_orders'))
        i = safe_float(inv.get('i_intransit'))
        l = safe_float(inv.get('l_prev_inventory'))
        m = safe_float(inv.get('m_prev_orders'))
        n = safe_float(inv.get('n_prev_intransit'))

        j = round(g - h + i, 1)
        o = round(l - m + n, 1)
        p_diff = round(j - o, 1)

        sales_2025 = safe_float(p.get('sales_2025'))
        total_shipped = safe_float(p.get('total_shipped'))
        status = calculate_status(j, sales_2025, p_diff, i, weeks_elapsed)

        weekly_rate_w = round(sales_2025 / weeks_elapsed, 1) if weeks_elapsed > 0 and sales_2025 > 0 else 0
        weeks_left_w = round(j / weekly_rate_w, 1) if weekly_rate_w > 0 else None
        weeks_left_p = round(j / abs(p_diff), 1) if p_diff and p_diff < 0 else None

        total_produced = safe_float(p.get('total_produced'))
        huiyang_inv = safe_float(p.get('huiyang_inv'))
        indonesia_inv = safe_float(p.get('indonesia_inv'))
        myanmar_inv = safe_float(p.get('myanmar_inv'))

        FOUR_MONTHS = 17
        prod_status = 'none'
        if j < 0 and abs(j) < total_produced:
            prod_status = 'judge'
        elif weeks_left_w is not None and weeks_left_w <= FOUR_MONTHS:
            if total_produced < 20:
                prod_status = 'produce'
            elif total_produced > 0:
                prod_status = 'available'

        merged.append({
            'item_code': code,
            'name': p.get('name', ''),
            'english_name': p.get('english_name', ''),
            'discontinued': p.get('discontinued', ''),
            'production_unit': p.get('production_unit', ''),
            'factory_inventory': safe_float(p.get('factory_inventory')),
            'shipping_warehouse': safe_float(p.get('shipping_warehouse')),
            'sales_2025': sales_2025,
            'total_shipped': total_shipped,
            'notes': p.get('notes', ''),
            'notes2': p.get('notes2', ''),
            'directive': p.get('directive', ''),
            'pl_mold': p.get('pl_mold', ''),
            'total_produced': total_produced,
            'huiyang_inv': huiyang_inv,
            'indonesia_inv': indonesia_inv,
            'myanmar_inv': myanmar_inv,
            'factory_unshipped': safe_float(p.get('factory_unshipped')),
            'website_on': p.get('website_on', ''),
            'web_note': p.get('web_note', ''),
            'g': g, 'h': h, 'i': i,
            'j': j, 'l': l, 'm': m, 'n': n, 'o': o, 'p': p_diff,
            'status': status,
            'prod_status': prod_status,
            'weekly_rate_w': weekly_rate_w,
            'weeks_left_w': weeks_left_w,
            'weeks_left_p': weeks_left_p,
        })
    return merged


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/products')
def api_products():
    status_filter = request.args.get('status')
    prod_filter = request.args.get('prod')
    search = request.args.get('search', '').lower()
    sort_by = request.args.get('sort', 'j')
    sort_dir = request.args.get('dir', 'asc')

    data = get_merged_data()

    if status_filter:
        data = [d for d in data if d['status'] == status_filter]
    if prod_filter:
        data = [d for d in data if d['prod_status'] == prod_filter]
    if search:
        data = [d for d in data if search in d['item_code'].lower() or search in d['name'].lower()]

    if sort_by in ('j', 'p', 'sales_2025', 'total_shipped', 'g', 'h', 'i', 'item_code', 'name', 'total_produced'):
        reverse = sort_dir == 'desc'
        data.sort(key=lambda x: (x.get(sort_by) is None, x.get(sort_by, '')), reverse=reverse)

    return jsonify({'products': data, 'total': len(data)})


@app.route('/api/stats')
def api_stats():
    data = get_merged_data()
    red = sum(1 for d in data if d['status'] == 'red')
    yellow = sum(1 for d in data if d['status'] == 'yellow')
    green = sum(1 for d in data if d['status'] == 'green')

    containers = read_csv(CONTAINERS_CSV)
    in_transit = sum(1 for c in containers if c.get('status') == 'in_transit')

    top_sellers = sorted([d for d in data if d['sales_2025'] > 0],
                          key=lambda x: x['sales_2025'], reverse=True)[:20]
    slow_sellers = sorted([d for d in data if d['sales_2025'] > 0],
                           key=lambda x: x['sales_2025'])[:20]

    biggest_drops = sorted([d for d in data if d['p'] < 0],
                            key=lambda x: x['p'])[:10]

    last_import = get_last_import()

    judge = sum(1 for d in data if d['prod_status'] == 'judge')
    available = sum(1 for d in data if d['prod_status'] == 'available')
    produce = sum(1 for d in data if d['prod_status'] == 'produce')

    return jsonify({
        'total_products': len(data),
        'red': red, 'yellow': yellow, 'green': green,
        'judge': judge, 'available': available, 'produce': produce,
        'in_transit_containers': in_transit,
        'last_import': last_import,
        'weeks_elapsed': get_weeks_elapsed(),
        'top_sellers': top_sellers,
        'slow_sellers': slow_sellers,
        'biggest_drops': biggest_drops,
    })


@app.route('/api/import', methods=['POST'])
def api_import():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({'error': 'No filename'}), 400

    suffix = os.path.splitext(file.filename)[1].lower()

    import_data = {}

    if suffix == '.xls':
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            file.save(tmp.name)
            tmp_path = tmp.name
        try:
            wb = xlrd.open_workbook(tmp_path)
            ws = wb.sheet_by_index(0)
            # Detect format: check if B has data (0812 format: A/B/C/D) or B is empty (Sheet1 format: A/C/E/G)
            has_col_b = False
            for r in range(1, min(20, ws.nrows)):
                if ws.cell_value(r, 1) != '' and ws.cell_value(r, 1) is not None:
                    has_col_b = True
                    break
            for r in range(1, ws.nrows):
                item = ws.cell_value(r, 0)
                if not item:
                    continue
                item = str(item).strip()
                if len(item) < 3:
                    continue
                if has_col_b:
                    po = ws.cell_value(r, 1)   # B
                    so = ws.cell_value(r, 2)   # C
                    on_hand = ws.cell_value(r, 3)  # D
                else:
                    po = ws.cell_value(r, 2)   # C
                    so = ws.cell_value(r, 4)   # E
                    on_hand = ws.cell_value(r, 6)  # G
                import_data[item] = {
                    'g': safe_float(on_hand),
                    'h': safe_float(so),
                    'i': safe_float(po),
                }
        finally:
            os.unlink(tmp_path)
    elif suffix == '.xlsx':
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            file.save(tmp.name)
            tmp_path = tmp.name
        try:
            wb = load_workbook(tmp_path, data_only=True)
            ws = wb.active
            has_col_b = False
            for r in range(2, min(20, ws.max_row + 1)):
                if ws.cell(row=r, column=2).value is not None:
                    has_col_b = True
                    break
            for r in range(2, ws.max_row + 1):
                item = ws.cell(row=r, column=1).value
                if not item:
                    continue
                item = str(item).strip()
                if len(item) < 3:
                    continue
                if has_col_b:
                    po = ws.cell(row=r, column=2).value   # B
                    so = ws.cell(row=r, column=3).value   # C
                    on_hand = ws.cell(row=r, column=4).value  # D
                else:
                    po = ws.cell(row=r, column=3).value   # C
                    so = ws.cell(row=r, column=5).value   # E
                    on_hand = ws.cell(row=r, column=7).value  # G
                import_data[item] = {
                    'g': safe_float(on_hand),
                    'h': safe_float(so),
                    'i': safe_float(po),
                }
        finally:
            os.unlink(tmp_path)
    elif suffix == '.csv':
        stream = file.read().decode('utf-8-sig').splitlines()
        reader = csv.DictReader(stream)
        headers_lower = {k.lower().strip(): k for k in (reader.fieldnames or [])}
        for row in reader:
            item = (row.get('Item') or row.get('item') or row.get('A') or '').strip()
            if not item or len(item) < 3:
                continue
            po = row.get('Quantity On Purchase Order') or row.get('B') or row.get('C') or 0
            so = row.get('Quantity On Sales Order') or row.get('C') or row.get('E') or 0
            on_hand = row.get('Quantity On Hand') or row.get('D') or row.get('G') or 0
            import_data[item] = {
                'g': safe_float(on_hand),
                'h': safe_float(so),
                'i': safe_float(po),
            }
    else:
        return jsonify({'error': f'Unsupported format: {suffix}'}), 400

    if not import_data:
        return jsonify({'error': 'No valid data found in file'}), 400

    products = read_csv(PRODUCTS_CSV)
    inventory = read_csv(INVENTORY_CSV)
    inv_map = {r['item_code']: r for r in inventory}
    prod_map = {p['item_code']: p for p in products}

    new_items = 0
    updated_items = 0

    for item_code, new_vals in import_data.items():
        if item_code in inv_map:
            old = inv_map[item_code]
            old['l_prev_inventory'] = old.get('g_inventory', '0')
            old['m_prev_orders'] = old.get('h_orders', '0')
            old['n_prev_intransit'] = old.get('i_intransit', '0')
            old['g_inventory'] = str(new_vals['g'])
            old['h_orders'] = str(new_vals['h'])
            old['i_intransit'] = str(new_vals['i'])
            updated_items += 1
        else:
            inv_map[item_code] = {
                'item_code': item_code,
                'g_inventory': str(new_vals['g']),
                'h_orders': str(new_vals['h']),
                'i_intransit': str(new_vals['i']),
                'l_prev_inventory': '0',
                'm_prev_orders': '0',
                'n_prev_intransit': '0',
            }
            inventory.append(inv_map[item_code])
            new_items += 1

        if item_code not in prod_map:
            products.append({
                'item_code': item_code, 'name': '', 'english_name': '',
                'discontinued': '', 'production_unit': '',
                'factory_inventory': '0', 'shipping_warehouse': '0',
                'sales_2025': '0', 'total_shipped': '0',
                'notes': '', 'notes2': '', 'directive': '', 'pl_mold': '',
                'total_produced': '0', 'huiyang_inv': '0',
                'indonesia_inv': '0', 'myanmar_inv': '0',
                'factory_unshipped': '0', 'website_on': '', 'web_note': '',
            })
            prod_map[item_code] = products[-1]

    inv_fields = ['item_code', 'g_inventory', 'h_orders', 'i_intransit',
                  'l_prev_inventory', 'm_prev_orders', 'n_prev_intransit']
    write_csv(INVENTORY_CSV, inventory, inv_fields)

    prod_fields = ['item_code', 'name', 'english_name', 'discontinued',
                   'production_unit', 'factory_inventory', 'shipping_warehouse',
                   'sales_2025', 'total_shipped', 'notes', 'notes2', 'directive',
                   'pl_mold', 'total_produced', 'huiyang_inv', 'indonesia_inv',
                   'myanmar_inv', 'factory_unshipped', 'website_on', 'web_note']
    write_csv(PRODUCTS_CSV, products, prod_fields)

    today_str = date.today().isoformat()
    snapshot_path = os.path.join(SNAPSHOT_DIR, f'{today_str}.csv')
    with open(snapshot_path, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.DictWriter(f, fieldnames=['item_code', 'g_inventory', 'h_orders', 'i_intransit'])
        w.writeheader()
        for item_code, vals in import_data.items():
            w.writerow({'item_code': item_code, 'g_inventory': vals['g'],
                        'h_orders': vals['h'], 'i_intransit': vals['i']})

    with open(LAST_IMPORT_JSON, 'w', encoding='utf-8') as f:
        json.dump({
            'date': today_str,
            'filename': file.filename,
            'item_count': len(import_data),
            'new_items': new_items,
            'updated_items': updated_items,
        }, f, ensure_ascii=False, indent=2)

    return jsonify({
        'success': True,
        'date': today_str,
        'filename': file.filename,
        'total_items': len(import_data),
        'new_items': new_items,
        'updated_items': updated_items,
    })


@app.route('/api/containers')
def api_containers():
    containers = read_csv(CONTAINERS_CSV)
    return jsonify({'containers': containers})


@app.route('/api/containers', methods=['POST'])
def api_add_container():
    data = request.json
    containers = read_csv(CONTAINERS_CSV)
    cid = f"{data.get('route','')}_{data.get('ship_date','')}"
    containers = [c for c in containers if c.get('id') != cid]
    containers.append({
        'id': cid,
        'ship_date': data.get('ship_date', ''),
        'eta': data.get('eta', ''),
        'route': data.get('route', ''),
        'status': 'in_transit',
    })
    write_csv(CONTAINERS_CSV, containers, ['id', 'ship_date', 'eta', 'route', 'status'])
    return jsonify({'success': True})


@app.route('/api/containers/<cid>', methods=['DELETE'])
def api_delete_container(cid):
    containers = read_csv(CONTAINERS_CSV)
    containers = [c for c in containers if c.get('id') != cid]
    write_csv(CONTAINERS_CSV, containers, ['id', 'ship_date', 'eta', 'route', 'status'])
    return jsonify({'success': True})


@app.route('/api/snapshots')
def api_snapshots():
    snaps = []
    if os.path.exists(SNAPSHOT_DIR):
        for fname in sorted(os.listdir(SNAPSHOT_DIR)):
            if fname.endswith('.csv'):
                snaps.append(fname.replace('.csv', ''))
    return jsonify({'snapshots': snaps})


@app.route('/api/export')
def api_export():
    data = get_merged_data()
    fields = ['item_code', 'name', 'english_name', 'g', 'h', 'i', 'j', 'l', 'm', 'n', 'o', 'p',
              'sales_2025', 'total_shipped', 'status', 'prod_status', 'production_unit',
              'factory_inventory', 'shipping_warehouse', 'notes', 'directive',
              'total_produced', 'huiyang_inv', 'indonesia_inv', 'myanmar_inv',
              'factory_unshipped', 'website_on', 'web_note']
    path = os.path.join(DATA_DIR, 'export_temp.csv')
    with open(path, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction='ignore')
        w.writeheader()
        w.writerows(data)
    return send_file(path, as_attachment=True, download_name='inventory_export.csv')


if __name__ == '__main__':
    app.run(debug=True, port=5000)
